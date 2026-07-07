from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path

import openpyxl


WORKBOOK_PATH = Path(r"C:\Users\hoffm\Desktop\Desktop\Toms\Familienplaner.xlsx")
OUTPUT_PATH = Path("prototype/data/seed-data.js")
TARGET_YEAR = date.today().year

DISPLAY_TITLE_OVERRIDES = {
    "drinks": "Getraenke aus Keller holen",
    "laundry-wash": "Waesche waschen + trocknen",
    "laundry-fold": "Waesche buegeln + zusammen legen",
    "own-room": "Eigene Zimmer reinigen",
}

MEMBER_COLORS = {
    "A": "#2563eb",
    "C": "#0f766e",
    "D": "#c2410c",
    "S": "#7c3aed",
    "T": "#be123c",
}

MEMBER_ROLES = {
    "Aaron": "child",
    "Cedrik": "child",
    "Danielle": "adult",
    "Sonja": "adult",
    "Tom": "adult",
}

SECTION_BY_ROW = {
    "daily": range(4, 9),
    "twice_weekly": range(11, 13),
    "weekly": range(15, 24),
    "long_term": range(26, 30),
}

# The weekly sheets store task labels as drawings/text boxes, but the schedule
# cells still follow this row order.
WEEK_SHEET_ROW_TASK_KEYS = {
    2: "compost",
    3: "dishwasher",
    4: "cook",
    5: "kitchen",
    6: "flowers",
    7: "shopping",
    8: "drinks",
    9: "trash",
    10: "paper-glass",
    11: "laundry-wash",
    12: "laundry-fold",
    13: "vacuum",
    14: "mop",
    15: "dust",
    16: "bathroom",
    17: "own-room",
    18: "descale",
    19: "bmw",
    20: "astra",
    21: "windows",
}

TASK_KEY_BY_TITLE_HINT = {
    "Kompost": "compost",
    "Geschirr": "dishwasher",
    "Essen zubereiten": "cook",
    "Kueche reinigen": "kitchen",
    "Küche reinigen": "kitchen",
    "Blumen": "flowers",
    "Getraenke": "drinks",
    "Getränke": "drinks",
    "einkaufen": "shopping",
    "Einkaufen": "shopping",
    "Restmuell": "trash",
    "Restmüll": "trash",
    "Altpapier": "paper-glass",
    "waschen": "laundry-wash",
    "trocknen": "laundry-wash",
    "zusammen": "laundry-fold",
    "buegeln": "laundry-fold",
    "bügeln": "laundry-fold",
    "Staubsaugen": "vacuum",
    "Nass": "mop",
    "Abstauben": "dust",
    "Baeder": "bathroom",
    "Bäder": "bathroom",
    "eigene": "own-room",
    "Entkalken": "descale",
    "BMW": "bmw",
    "Astra": "astra",
    "Fenster": "windows",
}

DAY_BY_COLUMN = {
    2: "Montag",
    3: "Dienstag",
    4: "Mittwoch",
    5: "Donnerstag",
    6: "Freitag",
    7: "Samstag",
    8: "Sonntag",
}

DAY_INDEX_BY_NAME = {
    "Montag": 1,
    "Dienstag": 2,
    "Mittwoch": 3,
    "Donnerstag": 4,
    "Freitag": 5,
    "Samstag": 6,
    "Sonntag": 7,
}

SCHEDULE_RULES = {
    "compost": {"type": "daily", "days": list(DAY_BY_COLUMN.values())},
    "dishwasher": {"type": "daily", "days": list(DAY_BY_COLUMN.values())},
    "cook": {"type": "daily", "days": list(DAY_BY_COLUMN.values())},
    "kitchen": {"type": "daily", "days": list(DAY_BY_COLUMN.values())},
    "flowers": {
        "type": "seasonal_daily",
        "days": list(DAY_BY_COLUMN.values()),
        "seasonStart": "06-01",
        "seasonEnd": "08-31",
    },
    "drinks": {"type": "weekly_days", "days": ["Dienstag", "Freitag"]},
    "shopping": {"type": "weekly_days", "days": ["Dienstag", "Samstag"]},
    "trash": {"type": "weekly_days", "days": ["Samstag"]},
    "paper-glass": {"type": "weekly_days", "days": ["Samstag"]},
    "laundry-wash": {"type": "weekly_days", "days": ["Samstag"]},
    "laundry-fold": {"type": "weekly_days", "days": ["Samstag"]},
    "vacuum": {"type": "weekly_days", "days": ["Samstag"]},
    "mop": {"type": "weekly_days", "days": ["Samstag"]},
    "dust": {"type": "weekly_days", "days": ["Samstag"]},
    "bathroom": {"type": "weekly_days", "days": ["Samstag"]},
    "own-room": {"type": "weekly_days", "days": ["Samstag"]},
    "descale": {"type": "interval_days", "days": ["Samstag"], "intervalDays": 28},
    "bmw": {"type": "interval_days", "days": ["Samstag"], "intervalDays": 56},
    "astra": {"type": "interval_days", "days": ["Samstag"], "intervalDays": 56, "offsetDays": 28},
    "windows": {"type": "interval_days", "days": ["Samstag"], "intervalDays": 175},
}


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def make_id(value: str) -> str:
    text = value.lower()
    replacements = {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }
    for source, replacement in replacements.items():
        text = text.replace(source, replacement)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "item"


def task_key_for_title(title: str) -> str:
    for hint, key in TASK_KEY_BY_TITLE_HINT.items():
        if hint in title:
            return key
    return make_id(title)


def category_for_row(row_number: int) -> str:
    for category, rows in SECTION_BY_ROW.items():
        if row_number in rows:
            return category
    return "custom"


def clean_task_title(raw_title: str) -> str:
    title = raw_title.lstrip("#").strip()
    replacements = {
        "anschmeißen": "anschmeissen",
        "gießen": "giessen",
        "Getränke": "Getraenke",
        "Küche": "Kueche",
        "Bäder": "Baeder",
        "Wäsche": "Waesche",
        "bügeln": "buegeln",
    }
    for source, replacement in replacements.items():
        title = title.replace(source, replacement)
    return title


def extract_members(ws):
    members = []
    for row in range(38, 43):
        name = normalize_text(ws.cell(row, 1).value)
        short_code = normalize_text(ws.cell(row, 2).value)
        if not name or not short_code:
            continue
        members.append(
            {
                "id": make_id(name),
                "name": name,
                "shortCode": short_code,
                "color": MEMBER_COLORS.get(short_code, "#0f766e"),
                "role": MEMBER_ROLES.get(name, "other"),
            }
        )
    return members


def extract_tasks(ws):
    tasks = []
    seen = set()
    for row in range(1, ws.max_row + 1):
        raw_title = normalize_text(ws.cell(row, 1).value)
        if not raw_title.startswith("#"):
            continue
        title = clean_task_title(raw_title)
        key = task_key_for_title(title)
        if key in seen:
            continue
        seen.add(key)
        title = DISPLAY_TITLE_OVERRIDES.get(key, title)
        task = {
            "id": key,
            "title": title,
            "category": category_for_row(row),
            "effortUnits": float(ws.cell(row, 2).value or 0),
            "source": "excel",
        }
        if "28 Tage" in raw_title:
            task["intervalDays"] = 28
        elif "56 Tage" in raw_title:
            task["intervalDays"] = 56
        elif "175 Tage" in raw_title:
            task["intervalDays"] = 175
        tasks.append(task)
    return tasks


def iso_weeks_in_year(year):
    return date(year, 12, 28).isocalendar().week


def first_day_for_weekday(year, day_name):
    day_index = DAY_INDEX_BY_NAME[day_name]
    first = date.fromisocalendar(year, 1, day_index)
    if first.year < year:
        first = date.fromisocalendar(year, 2, day_index)
    return first


def is_in_season(task_date, rule):
    start_month, start_day = [int(part) for part in rule["seasonStart"].split("-")]
    end_month, end_day = [int(part) for part in rule["seasonEnd"].split("-")]
    start = date(task_date.year, start_month, start_day)
    end = date(task_date.year, end_month, end_day)
    return start <= task_date <= end


def is_due_on_date(task_id, task_date, day_name):
    rule = SCHEDULE_RULES.get(task_id)
    if not rule or day_name not in rule["days"]:
        return False

    rule_type = rule["type"]
    if rule_type in {"daily", "weekly_days"}:
        return True

    if rule_type == "seasonal_daily":
        return is_in_season(task_date, rule)

    if rule_type == "interval_days":
        start = first_day_for_weekday(task_date.year, rule["days"][0])
        offset_days = rule.get("offsetDays", 0)
        start = start.replace(day=start.day)  # keep this as a date, not datetime
        delta = (task_date - start).days - offset_days
        return delta >= 0 and delta % rule["intervalDays"] == 0

    return False


def extract_imported_assignments(wb, members_by_code, tasks_by_id):
    assignments_by_week = {}
    source_weeks = []
    for sheet_name in wb.sheetnames:
        if not re.fullmatch(r"KW\d+", sheet_name):
            continue
        week = int(sheet_name[2:])
        source_weeks.append(week)
        assignments_by_week[week] = []
        ws = wb[sheet_name]
        for row, task_id in WEEK_SHEET_ROW_TASK_KEYS.items():
            if task_id not in tasks_by_id:
                continue
            for col, day in DAY_BY_COLUMN.items():
                value = normalize_text(ws.cell(row, col).value)
                if not value:
                    continue
                for short_code in value:
                    member_id = members_by_code.get(short_code)
                    if not member_id:
                        continue
                    assignments_by_week[week].append(
                        {
                            "taskId": task_id,
                            "memberId": member_id,
                            "day": day,
                            "dayIndex": DAY_INDEX_BY_NAME[day],
                            "status": "open",
                        }
                    )
    return sorted(source_weeks), assignments_by_week


def source_members_for(assignments_by_week, source_weeks, source_week, task_id, day):
    members = [
        assignment["memberId"]
        for assignment in assignments_by_week.get(source_week, [])
        if assignment["taskId"] == task_id and assignment["day"] == day
    ]
    if members:
        return members

    for fallback_week in source_weeks:
        members = [
            assignment["memberId"]
            for assignment in assignments_by_week.get(fallback_week, [])
            if assignment["taskId"] == task_id and assignment["day"] == day
        ]
        if members:
            return members

    return []


def build_schedule_rules(tasks):
    rules = []
    for task in tasks:
        rule = SCHEDULE_RULES.get(task["id"])
        if not rule:
            continue
        rules.append(
            {
                "id": f"rule-{task['id']}",
                "taskId": task["id"],
                **rule,
            }
        )
    return rules


def build_year_assignments(source_weeks, assignments_by_week, tasks, year):
    assignments = []
    week_count = iso_weeks_in_year(year)
    task_ids = [task["id"] for task in tasks]
    for week in range(1, week_count + 1):
        source_week = source_weeks[(week - 1) % len(source_weeks)]
        for day in DAY_BY_COLUMN.values():
            task_date = date.fromisocalendar(year, week, DAY_INDEX_BY_NAME[day])
            for task_id in task_ids:
                if not is_due_on_date(task_id, task_date, day):
                    continue
                members = source_members_for(assignments_by_week, source_weeks, source_week, task_id, day)
                for member_index, member_id in enumerate(members, start=1):
                    assignments.append(
                        {
                            "id": f"{year}-kw{week}-{task_id}-{make_id(day)}-{member_id}-{member_index}",
                            "year": year,
                            "week": week,
                            "sourceWeek": source_week,
                            "date": task_date.isoformat(),
                            "taskId": task_id,
                            "memberId": member_id,
                            "day": day,
                            "dayIndex": DAY_INDEX_BY_NAME[day],
                            "status": "open",
                            "generatedFromRule": f"rule-{task_id}",
                        }
                    )
    return list(range(1, week_count + 1)), assignments


def extract_meal_pattern(wb, members_by_code):
    if "Essensplan" not in wb.sheetnames:
        return []

    ws = wb["Essensplan"]
    meals = []
    current_week = None
    for row in range(2, ws.max_row + 1):
        week_value = ws.cell(row, 1).value
        if week_value:
            current_week = int(week_value)

        day = normalize_text(ws.cell(row, 2).value)
        title = normalize_text(ws.cell(row, 3).value)
        cook_code = normalize_text(ws.cell(row, 4).value)
        if not current_week or not day:
            continue

        meals.append(
            {
                "id": f"kw{current_week}-meal-{make_id(day)}",
                "week": current_week,
                "day": day,
                "title": title,
                "cookMemberId": members_by_code.get(cook_code) if cook_code else None,
            }
        )
    return meals


def build_year_meals(meal_pattern, year):
    by_week = {}
    for meal in meal_pattern:
        by_week.setdefault(meal["week"], []).append(meal)

    source_weeks = sorted(by_week)
    if not source_weeks:
        return []

    meals = []
    week_count = iso_weeks_in_year(year)
    for week in range(1, week_count + 1):
        source_week = source_weeks[(week - 1) % len(source_weeks)]
        for source_meal in by_week[source_week]:
            day_index = DAY_INDEX_BY_NAME.get(source_meal["day"])
            meal_date = date.fromisocalendar(year, week, day_index) if day_index else None
            meals.append(
                {
                    **source_meal,
                    "id": f"{year}-kw{week}-meal-{make_id(source_meal['day'])}",
                    "year": year,
                    "week": week,
                    "sourceWeek": source_week,
                    "date": meal_date.isoformat() if meal_date else None,
                }
            )
    return meals


def main():
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    overview = workbook["Übersichtsplaner"]

    members = extract_members(overview)
    members_by_code = {member["shortCode"]: member["id"] for member in members}

    tasks = extract_tasks(overview)
    tasks_by_id = {task["id"]: task for task in tasks}

    source_weeks, imported_assignments_by_week = extract_imported_assignments(workbook, members_by_code, tasks_by_id)
    schedule_rules = build_schedule_rules(tasks)
    weeks, assignments = build_year_assignments(source_weeks, imported_assignments_by_week, tasks, TARGET_YEAR)
    meal_pattern = extract_meal_pattern(workbook, members_by_code)
    meals = build_year_meals(meal_pattern, TARGET_YEAR)

    data = {
        "family": {
            "id": "family-hoffmann",
            "name": "Familie Hoffmann",
            "week": date.today().isocalendar().week,
            "year": TARGET_YEAR,
            "availableWeeks": weeks,
            "weeklyTargetPerMember": 14.69,
            "sourceWeeks": source_weeks,
        },
        "members": members,
        "taskTemplates": tasks,
        "scheduleRules": schedule_rules,
        "assignments": assignments,
        "meals": meals,
        "source": {
            "file": str(WORKBOOK_PATH),
            "notes": [
                "Aufgaben und Einheiten stammen aus dem Blatt Uebersichtsplaner.",
                "Sichtbare Wochenblatt-Bezeichnungen werden fuer einzelne Aufgaben bevorzugt.",
                f"Wochenzuweisungen stammen aus KW{source_weeks[0]} bis KW{source_weeks[-1]} und werden fuer {TARGET_YEAR} als Jahresmuster fortgeschrieben.",
                "Essensplan stammt aus dem Blatt Essensplan und wird als Jahresmuster fortgeschrieben.",
            ],
        },
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    encoded = json.dumps(data, ensure_ascii=False, indent=2)
    OUTPUT_PATH.write_text(f"window.seedData = {encoded};\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "members": len(members),
                "tasks": len(tasks),
                "rules": len(schedule_rules),
                "weeks": len(weeks),
                "sourceWeeks": len(source_weeks),
                "assignments": len(assignments),
                "meals": len(meals),
                "output": str(OUTPUT_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
